import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.services.prediction_components import AMSTERDAM
# from src.pages.pivot import get_pivot_data

# Brand colors (from theme.py, as hex without #)
_INK       = "16282B"
_GREEN     = "3E6B4F"
_BROWN     = "7A4B2A"
_CARD_BG   = "FBF5EA"
_DIVIDER   = "E7DDC8"
_WHITE     = "FFFFFF"
_MUTED     = "5B5247"

def export_pivot_to_xlsx(path: str = "wc_predictions.xlsx") -> str:
    from src.pages.pivot import get_pivot_data
    matches, rows = get_pivot_data()  # your existing function

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(hex_color=_INK, bold=False, size=10):
        return Font(name="Arial", color=hex_color, bold=bold, size=size)

    def _border():
        side = Side(style="thin", color=_DIVIDER)
        return Border(left=side, right=side, top=side, bottom=side)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Header row ───────────────────────────────────────────────────────────
    headers = ["User", "Total pts"]
    for m in matches:
        kickoff = (
            m.match_date.astimezone(AMSTERDAM).strftime("%d/%m %H:%M")
            if m.match_date else "—"
        )
        headers.append(
            f"{m.home_team[:3].upper()} v {m.away_team[:3].upper()}\n{kickoff}\n{m.stage or ''}"
        )

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(name="Arial", color=_WHITE, bold=True, size=10)
        cell.fill      = _fill(_INK)
        cell.alignment = center
        cell.border    = _border()

    ws.row_dimensions[1].height = 56

    # ── Data rows ────────────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows, start=2):
        bg = _CARD_BG if r_idx % 2 == 0 else _WHITE

        # Login
        c = ws.cell(row=r_idx, column=1, value=row["login"])
        c.font = _font(bold=True); c.fill = _fill(bg)
        c.alignment = left; c.border = _border()

        # Total points
        c = ws.cell(row=r_idx, column=2, value=row["total"])
        c.font = Font(name="Arial", color=_GREEN, bold=True, size=10)
        c.fill = _fill(bg); c.alignment = center; c.border = _border()

        # One cell per match
        for m_idx, cell_data in enumerate(row["cells"], start=3):
            c = ws.cell(row=r_idx, column=m_idx)
            c.fill = _fill(bg); c.alignment = center; c.border = _border()

            if cell_data["done"]:
                pts = cell_data["points"]
                c.value = f"{cell_data['pred']}\n({pts} pts)"
                color = _GREEN if pts > 0 else _BROWN
                c.font = Font(name="Arial", color=color, size=10)
            else:
                c.value = "—"
                c.font = _font(_DIVIDER)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18  # login
    ws.column_dimensions["B"].width = 10  # total
    for i in range(len(matches)):
        ws.column_dimensions[get_column_letter(i + 3)].width = 13

    # ── Freeze panes: keep login + header visible while scrolling ────────────
    ws.freeze_panes = "C2"

    wb.save(path)
    return path